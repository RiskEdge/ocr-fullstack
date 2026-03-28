"""
Seed the `master_items` table from a CSV or XLSX file.

The table must already exist (run migrations/create_master_items.sql first).

Usage:
    python seed_master_data.py <file_path> [--sheet <sheet_name>] [--clear]

Arguments:
    file_path       Path to a .csv or .xlsx file containing master data.

Options:
    --sheet         Sheet name to read from an XLSX file.
                    Defaults to the first sheet whose name contains "master"
                    (case-insensitive), or the first sheet if none match.
    --clear         Delete ALL existing rows before inserting.
                    Default behaviour is upsert (insert or update on plu_code conflict).

Expected columns (case-insensitive, spaces/underscores ignored):
    SkuCode, PluCode, Priority, SkuDesc, EanCode, CostPrice, MRP, Tax%

Example:
    python seed_master_data.py "Sample data.xlsx"
    python seed_master_data.py master.csv --clear
"""

import argparse
import csv
import io
import os
import sys
import dotenv

dotenv.load_dotenv()

from app.db import get_supabase


# ---------------------------------------------------------------------------
# Column-name normalisation
# ---------------------------------------------------------------------------

# Maps normalised header → canonical field name
_COLUMN_MAP: dict[str, str] = {
    "skucode":    "sku_code",
    "plucode":    "plu_code",
    "priority":   "priority",
    "skudesc":    "sku_desc",
    "eancode":    "ean_code",
    "ean":        "ean_code",
    "barcode":    "ean_code",
    "costprice":  "cost_price",
    "mrp":        "mrp",
    "tax%":       "tax_pct",
    "tax":        "tax_pct",
    "gst%":       "tax_pct",
}


def _norm(h: object) -> str:
    """Lowercase, strip spaces/underscores — 'Cost Price' → 'costprice'."""
    return str(h).strip().lower().replace(" ", "").replace("_", "") if h is not None else ""


def _to_float(val: object) -> float | None:
    try:
        return float(val)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _ean_str(val: object) -> str | None:
    """Convert 8901399000591.0 → '8901399000591'."""
    if val is None:
        return None
    try:
        return str(int(float(str(val))))
    except (ValueError, TypeError):
        s = str(val).strip()
        return s or None


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_headers(raw_headers: list) -> dict[str, int]:
    """Return {canonical_field: column_index} for recognised headers."""
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        canonical = _COLUMN_MAP.get(_norm(h))
        if canonical and canonical not in col_index:
            col_index[canonical] = i
    return col_index


def _build_record(row: list | dict, col_index: dict[str, int]) -> dict | None:
    """Convert one raw row into a DB-ready dict. Returns None if EAN is missing."""
    def get(field: str):
        idx = col_index.get(field)
        if idx is None:
            return None
        val = row[idx] if isinstance(row, list) else list(row.values())[idx]
        return val

    ean = _ean_str(get("ean_code"))
    if not ean:
        return None

    plu_raw = get("plu_code")
    plu = _ean_str(plu_raw)   # PluCode is also numeric
    if not plu:
        return None

    return {
        "ean_code":   ean,
        "plu_code":   plu,
        "sku_code":   _ean_str(get("sku_code")),
        "priority":   int(_to_float(get("priority")) or 1),
        "sku_desc":   str(get("sku_desc")).strip() if get("sku_desc") is not None else None,
        "cost_price": _to_float(get("cost_price")),
        "mrp":        _to_float(get("mrp")),
        "tax_pct":    _to_float(get("tax_pct")),
    }


def read_xlsx(path: str, sheet: str | None) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            sys.exit(1)
        sheet_name = sheet
    else:
        sheet_name = next(
            (s for s in wb.sheetnames if "master" in s.lower()),
            wb.sheetnames[0],
        )

    print(f"  Reading sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    col_index = _parse_headers(list(all_rows[0]))
    _validate_col_index(col_index)

    records = []
    skipped = 0
    for row in all_rows[1:]:
        rec = _build_record(list(row), col_index)
        if rec:
            records.append(rec)
        else:
            skipped += 1

    if skipped:
        print(f"  Skipped {skipped} row(s) with missing EAN or PluCode.")
    return records


def read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        raw_headers = reader.fieldnames or []
        col_index = _parse_headers(raw_headers)
        _validate_col_index(col_index)

        records = []
        skipped = 0
        for row in reader:
            values = list(row.values())
            rec = _build_record(values, col_index)
            if rec:
                records.append(rec)
            else:
                skipped += 1

    if skipped:
        print(f"  Skipped {skipped} row(s) with missing EAN or PluCode.")
    return records


def _validate_col_index(col_index: dict[str, int]) -> None:
    missing = [f for f in ("ean_code", "plu_code") if f not in col_index]
    if missing:
        print(f"ERROR: Could not find required column(s): {missing}")
        print("       Check that your file has EanCode and PluCode columns.")
        sys.exit(1)


# ---------------------------------------------------------------------------
# DB operations
# ---------------------------------------------------------------------------

CHUNK = 500  # Supabase batch insert limit


def seed(records: list[dict], clear: bool) -> None:
    db = get_supabase()

    if clear:
        db.table("master_items").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        print("  Cleared existing master_items rows.")

    inserted = 0
    for i in range(0, len(records), CHUNK):
        chunk = records[i : i + CHUNK]
        db.table("master_items").upsert(chunk, on_conflict="plu_code").execute()
        inserted += len(chunk)
        print(f"  Upserted {inserted}/{len(records)} rows...", end="\r")

    print(f"  Upserted {inserted} rows.          ")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Seed master_items from CSV or XLSX.")
    parser.add_argument("file", help="Path to .csv or .xlsx master data file")
    parser.add_argument("--sheet", default=None, help="XLSX sheet name (optional)")
    parser.add_argument("--clear", action="store_true", help="Delete all rows before inserting")
    args = parser.parse_args()

    path: str = args.file
    if not os.path.exists(path):
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    ext = os.path.splitext(path)[1].lower()
    print(f"\nReading: {path}")

    if ext == ".xlsx":
        records = read_xlsx(path, args.sheet)
    elif ext == ".csv":
        records = read_csv(path)
    else:
        print(f"ERROR: Unsupported file type '{ext}'. Use .csv or .xlsx.")
        sys.exit(1)

    if not records:
        print("No valid records found. Nothing to seed.")
        sys.exit(0)

    print(f"  Parsed {len(records)} valid record(s).")
    seed(records, clear=args.clear)
    print("Done.\n")


if __name__ == "__main__":
    main()
